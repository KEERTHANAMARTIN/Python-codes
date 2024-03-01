import openpyxl
# go to the location of the xl path
book= openpyxl.load_workbook("C:\\Users\\admin\\Downloads\\input.xlsx")

# to open the xl sheet
sheet = book.active

# to identify the rows and columns in the xl sheet
cell = sheet.cell(row= 2, column =2)

# to print the values of the cell
print(cell.value)


# taking data's from 2 Excel sheets

from openpyxl import load_workbook

# load the workbook
workbook = load_workbook("C:\\Users\\admin\\Downloads\\PowerBi\\123.xlsx")

#extract data from sheet1
sheet1 = workbook['Sheet1']
data_sheet1 =[]

for ro in sheet1.iter_rows(values_only=True):
    data_sheet1.append(row)

#Extract data from sheet2
sheet2 = workbook['Sheet2']
data_sheet2=[]

for row in sheet2.iter_rows(values_only=True):
    data_sheet2.append(row)

#printing the data from sheet1
print("Data from sheet1:")
for row in data_sheet1:
    print(row)

#printing the data from sheet2
print("Data from Sheet2:")
for row in data_sheet2:
    print(row)