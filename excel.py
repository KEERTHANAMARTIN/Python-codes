# taking data's from 2 Excel sheets

from openpyxl import load_workbook

# load the workbook
workbook = load_workbook("C:\\Users\\admin\\Downloads\\Book2.xlsx")

#extract data from sheet1
sheet1 = workbook['Sheet1']
data_sheet1 =[]

for row in sheet1.iter_rows():
    data_row =[]
    for cell in row:
        data_row.append(cell.value)
    data_sheet1.append(data_row)


#Extract data from sheet2
sheet2 = workbook['Sheet2']
data_sheet2=[]

for row in sheet2.iter_rows():
    data_row =[]
    for cell in row:
        data_row.append(cell.value)
    data_sheet2.append(data_row)

#printing the data from sheet1
print("Data from sheet1:")
for row in data_sheet1:
    print(row)

#printing the data from sheet2
print("Data from Sheet2:")
for row in data_sheet2:
    print(row)